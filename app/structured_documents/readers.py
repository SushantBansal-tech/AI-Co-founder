import csv
from pathlib import Path


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def read_tabular_rows(path: str) -> list[dict[str, str]]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("XLSX ingestion requires openpyxl") from exc
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        try:
            headers = [_clean(value).lower() for value in next(values)]
        except StopIteration:
            return []
        rows = [dict(zip(headers, row)) for row in values]
    else:
        raise ValueError(f"Structured ingestion does not support {suffix}")

    normalized = []
    for row_number, row in enumerate(rows, start=2):
        clean = {_clean(key).lower(): _clean(value) for key, value in row.items() if key is not None}
        if any(clean.values()):
            clean["_row_number"] = str(row_number)
            normalized.append(clean)
    return normalized
